#!/usr/bin/env python3
"""
Build a multi-sheet XLSX of the SME Mart compliance catalog for Brian.

Data source is the hand-curated catalog in
.claude/proposals/sme-mart-cyberab-catalog.md (mirrored into the rows below
rather than parsed — catalog is small enough that explicit data is clearer).

Usage:
    python3 scripts/build-compliance-catalog-xlsx.py [output.xlsx]
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = sys.argv[1] if len(sys.argv) > 1 else ".claude/proposals/sme-mart-compliance-catalog.xlsx"

# ──────────────────────────────────────────────────────────────────────────
# DATA — Roles (ecosystem org and individual roles)
# ──────────────────────────────────────────────────────────────────────────

ROLES = [
    # (ID, Role, Ecosystem, Scope, Summary, Prerequisites, Fee/Renewal, Source)
    # CMMC
    ("CMMC.RPO", "Registered Practitioner Organization", "CMMC", "Org",
     "Consulting MSP — employs RPs; does not assess.", "Background check",
     "$6,000 app / $5,000 annual", "cyberab.org"),
    ("CMMC.RP", "Registered Practitioner", "CMMC", "Individual",
     "Entry-level CMMC implementation consultant.", "Training + exam",
     "$600 app+training+test / $500 annual", "cyberab.org"),
    ("CMMC.RPA", "Registered Practitioner Advanced", "CMMC", "Individual",
     "Advanced CMMC consultant.", "Active RP + eligibility",
     "$1,000 app+training+test / $750 annual", "cyberab.org"),
    ("CMMC.C3PAO", "CMMC Third-Party Assessment Organization", "CMMC", "Org",
     "Authorized to conduct Level 2 assessments. Employs CCPs/CCAs.",
     "Experian + DCSA FOCI + DCMA DIBCAC Level 2 assessment",
     "$6,000 app + $15,000 authorization", "cyberab.org"),
    ("CMMC.CCP", "Certified CMMC Professional", "CMMC", "Individual",
     "Level 1 assessor (cannot make final determinations).",
     "Degree in cyber/IT OR 2+ yrs; CompTIA A+; DoD CUI training; Tier 3",
     "$200 reg + $275 exam / $250 annual", "cyberab.org + isaca.org/cmmc"),
    ("CMMC.CCA", "Certified CMMC Assessor", "CMMC", "Individual",
     "Full Level 2 assessor.",
     "Active CCP; 3+ yrs cyber; 1+ yr assessment; 1x 8140.3 cert; Tier 3",
     "$50 reg + $350 exam / $500 annual", "cyberab.org + isaca.org/cmmc"),
    ("CMMC.LCCA", "Lead CMMC Certified Assessor", "CMMC", "Individual",
     "Senior assessor; leads assessment teams.",
     "Active CCA + enhanced DoD experience",
     "$100 reg / $100 annual", "cyberab.org"),
    ("CMMC.CCI", "CMMC Credentialed Instructor", "CMMC", "Individual",
     "Instructs CMMC courses for ATPs. Launching early 2026.",
     "Meet assessor reqs at level instructed; coming early 2026",
     "$2,500 app+training+test / annual", "cyberab.org + isaca.org/cmmc"),
    ("CMMC.ATP", "Approved Training Provider", "CMMC", "Org",
     "Vetted by CAICO to deliver CMMC training (formerly LTP).",
     "CAICO approval", "See LTP-Registration", "cyberab.org"),
    ("CMMC.APP", "Approved Publishing Partner", "CMMC", "Org",
     "Vetted by CAICO for curriculum. NOT accepting applications.",
     "Experian financial review", "$6,000 app / $5,000 annual", "cyberab.org"),
    ("CMMC.OSC", "Organization Seeking Certification", "CMMC", "Org",
     "DIB company pursuing CMMC cert (customer side).", "—", "—", "cyberab.org"),
    ("CMMC.CQAP", "CMMC Quality Assurance Professional", "CMMC", "Individual",
     "Cyber AB-trained; QA for assessment docs.", "Cyber AB training", "—",
     "cyberab.org (Terminology)"),
    # SCF
    ("SCF.RPO", "SCF Registered Provider Organization", "SCF", "Org",
     "Consults on SCF frameworks.",
     "Background check; min. 1 SCF Practitioner or Architect on staff",
     "$6,000 app / $5,000 annual", "cyberab.org"),
    ("SCF.3PAO", "SCF Third-Party Assessment Organization", "SCF", "Org",
     "Conducts SCF conformity assessments.",
     "Commercial background; procedural assessment; min. 2 SCF Assessors",
     "$5,000 app / $5,000 authorization", "cyberab.org"),
    ("SCF.APP", "SCF Authorized Platform Partner", "SCF", "Org",
     "Integrates SCF into GRC platforms.",
     "Background; SCF integration; min. 1 SCF Architect",
     "$10,000 app / $10,000 annual", "cyberab.org"),
    ("SCF.APO", "SCF Authorized Platform Organization", "SCF", "Org",
     "Cloud platform for SCF implementation (CSPs, cloud RM).",
     "Background; SCF Platform Assessment; min. 1 SCF Architect",
     "$6,000 app / $5,000 annual", "cyberab.org"),
    ("SCF.Practitioner", "SCF Practitioner", "SCF", "Individual",
     "Individual credential — staff min for SCF RPO.", "SCF Council managed",
     "—", "cyberab.org"),
    ("SCF.Architect", "SCF Architect", "SCF", "Individual",
     "Senior individual credential — staff min for SCF APO/APP.",
     "SCF Council managed", "—", "cyberab.org"),
    ("SCF.Assessor", "SCF Assessor", "SCF", "Individual",
     "Individual — min 2 required for SCF 3PAO.", "SCF Council managed", "—",
     "cyberab.org"),
    ("SCF.OSA", "Organization Seeking Assessment", "SCF", "Org",
     "Client pursuing SCF certification (customer side).", "—", "—",
     "cyberab.org"),
    # SCA
    ("SCA.3PAO", "SCA Third-Party Assessment Organization", "SCA", "Org",
     "Issues SCA CODE 1/2/3 certifications (typically via SCF 3PAO status).",
     "Cyber AB accreditation", "—", "cyberab.org"),
    ("SCA.SDO", "Secure Development Organization", "SCA", "Org",
     "Org pursuing CODE 1/2/3 certification (customer side).", "Audit + ROC",
     "—", "cyberab.org"),
    ("SCA.Practitioner", "SCA Practitioner", "SCA", "Individual",
     "SAICO-managed individual credential for secure coding.",
     "SAICO requirements", "—", "cyberab.org"),
    ("SCA.Architect", "SCA Architect", "SCA", "Individual",
     "SAICO-managed senior credential for secure software architecture.",
     "SAICO requirements", "—", "cyberab.org"),
    # FedRAMP
    ("FedRAMP.3PAO", "FedRAMP 3PAO", "FedRAMP", "Org",
     "Accredited by A2LA to assess Cloud Service Offerings.",
     "ISO/IEC 17020 accreditation via A2LA", "—", "fedramp.gov"),
    ("FedRAMP.CSO", "FedRAMP Cloud Service Offering", "FedRAMP", "Product",
     "Authorized or In-Process product on FedRAMP Marketplace.",
     "3PAO assessment + agency/JAB ATO", "—", "fedramp.gov"),
    # HITRUST
    ("HITRUST.AEAO", "HITRUST Authorized External Assessor Organization",
     "HITRUST", "Org",
     "Licensed to perform HITRUST assessments via MyCSF.",
     "Active HITRUST license", "—", "hitrustalliance.net"),
    ("HITRUST.CCSFP", "HITRUST Certified CSF Practitioner", "HITRUST",
     "Individual", "Required for HITRUST assessors.", "HITRUST Academy", "—",
     "hitrustalliance.net"),
    ("HITRUST.CHQP", "HITRUST Certified Quality Professional", "HITRUST",
     "Individual", "Senior QA credential.", "HITRUST Academy", "—",
     "hitrustalliance.net"),
    # PCI SSC — org-level
    ("PCI.QSA", "Qualified Security Assessor", "PCI", "Org",
     "Conducts PCI DSS assessments.", "PCI SSC training + exam", "—",
     "pcisecuritystandards.org"),
    ("PCI.ASV", "Approved Scanning Vendor", "PCI", "Org",
     "Performs external vulnerability scans.", "PCI SSC approval", "—",
     "pcisecuritystandards.org"),
    ("PCI.ISA", "Internal Security Assessor Sponsor", "PCI", "Org",
     "Sponsors in-house PCI staff.", "PCI SSC sponsorship", "—",
     "pcisecuritystandards.org"),
    ("PCI.PA-QSA", "Payment Application QSA", "PCI", "Org",
     "Assesses payment applications (PA-DSS successor scope).",
     "QSA status + PA qualification", "—", "pcisecuritystandards.org"),
    ("PCI.P2PE-Assessor", "P2PE Assessor", "PCI", "Org",
     "Assesses Point-to-Point Encryption solutions.",
     "P2PE qualification", "—", "pcisecuritystandards.org"),
    ("PCI.QPA", "Qualified PIN Assessor", "PCI", "Org",
     "PIN transaction security.", "PIN qualification", "—",
     "pcisecuritystandards.org"),
    ("PCI.CPSA", "Card Production Security Assessor", "PCI", "Org",
     "Physical/logical card production assessments.", "CPSA qualification",
     "—", "pcisecuritystandards.org"),
    ("PCI.3DS-Assessor", "3-D Secure Assessor", "PCI", "Org",
     "Assesses 3DS components.", "3DS qualification", "—",
     "pcisecuritystandards.org"),
    ("PCI.Secure-Software-Assessor", "Secure Software Assessor", "PCI", "Org",
     "Assesses software per Secure Software Standard.",
     "Secure Software Assessor qualification", "—",
     "pcisecuritystandards.org"),
    ("PCI.Secure-SLC-Assessor", "Secure SLC Assessor", "PCI", "Org",
     "Assesses Software Lifecycle.", "Secure SLC qualification", "—",
     "pcisecuritystandards.org"),
    ("PCI.QIR", "Qualified Integrator and Reseller", "PCI", "Org",
     "Installs/configures POS per PCI DSS.", "QIR qualification", "—",
     "pcisecuritystandards.org"),
    ("PCI.PFI", "PCI Forensic Investigator", "PCI", "Individual",
     "Investigates card-data breaches.", "Must work for QSA company", "—",
     "pcisecuritystandards.org"),
    ("PCI.PCIP", "PCI Professional", "PCI", "Individual",
     "Individual PCI DSS competency.", "PCI SSC exam", "—",
     "pcisecuritystandards.org"),
    # SOC / AICPA
    ("AICPA.CPA-Firm", "Licensed CPA Firm", "SOC 2", "Org",
     "Authority to sign SOC 2 opinion.", "State CPA license + AICPA peer review",
     "—", "aicpa-cima.com"),
    ("AICPA.CPA-Individual", "Certified Public Accountant", "SOC 2",
     "Individual", "State-licensed CPA; typical SOC auditor.",
     "State CPA exam + experience", "State-specific", "state CPA boards"),
    # CSA STAR
    ("CSA.STAR-L1-Self", "CSA STAR Level 1 Self-Assessment", "CSA STAR", "Org",
     "CAIQ submission (annual).", "CAIQ completion", "Free; annual updates",
     "cloudsecurityalliance.org"),
    ("CSA.STAR-L1-ValidAIted", "CSA STAR Valid-AI-ted", "CSA STAR", "Org",
     "AI-validated CAIQ.", "—", "$595 (free for Corp members)",
     "cloudsecurityalliance.org"),
    ("CSA.STAR-L2-Attestation", "CSA STAR Attestation", "CSA STAR", "Org",
     "SOC 2 + CCM.", "Pass SOC 2 audit", "1-year validity",
     "cloudsecurityalliance.org"),
    ("CSA.STAR-L2-Certification", "CSA STAR Certification", "CSA STAR", "Org",
     "ISO/IEC 27001 + CCM.", "Hold ISO 27001", "3-year validity",
     "cloudsecurityalliance.org"),
    ("CSA.C-STAR", "C-STAR (Greater China)", "CSA STAR", "Org",
     "GB/T 22080 + CCM.", "GB/T 22080 cert", "3-year validity",
     "cloudsecurityalliance.org"),
    ("CSA.STAR-AI-L1", "CSA STAR for AI Level 1", "CSA STAR", "Org",
     "AI-CAIQ self-assessment.", "—", "—", "cloudsecurityalliance.org"),
    ("CSA.STAR-AI-L2", "CSA STAR for AI Level 2", "CSA STAR", "Org",
     "AI-CAIQ + ISO/IEC 42001.", "ISO 42001 cert", "—",
     "cloudsecurityalliance.org"),
    # CREST
    ("CREST.Pathway", "CREST Pathway Member", "CREST", "Org",
     "CREST member tier (entry).", "CREST membership", "—",
     "crest-approved.org"),
    ("CREST.Approved-Cyber-Assurance", "CREST Approved — Cyber Assurance",
     "CREST", "Org", "Cyber Assurance services.", "CREST approval", "—",
     "crest-approved.org"),
    ("CREST.Approved-IR", "CREST Approved — Incident Management", "CREST",
     "Org", "IR services.", "CREST approval", "—", "crest-approved.org"),
    ("CREST.Approved-Red-Team", "CREST Approved — Red Teaming", "CREST",
     "Org", "Red team services.", "CREST approval", "—", "crest-approved.org"),
    ("CREST.Approved-SOC", "CREST Approved — Security Ops Centre", "CREST",
     "Org", "SOC services.", "CREST approval", "—", "crest-approved.org"),
    ("CREST.Approved-Testing", "CREST Approved — Security Testing", "CREST",
     "Org", "Pentest services.", "CREST approval", "—", "crest-approved.org"),
    ("CREST.Approved-TI", "CREST Approved — Threat Intelligence", "CREST",
     "Org", "Threat intel services.", "CREST approval", "—",
     "crest-approved.org"),
]

# ──────────────────────────────────────────────────────────────────────────
# DATA — Certifications (individual credentials from all issuers)
# ──────────────────────────────────────────────────────────────────────────

CERTS = [
    # (ID, Certification, Issuer, Ecosystem, Proficiency, Scope, Notes, Source)
    # DoD 8140.3 Career Pathway 612 — Intermediate
    ("8140-612.CASP+", "SecurityX / CASP+", "CompTIA", "DoD 8140.3 / 612",
     "Intermediate", "Individual", "Accepted for CCA eligibility",
     "cyberab.org"),
    ("8140-612.CGRC", "CGRC / CAP", "(ISC)²", "DoD 8140.3 / 612",
     "Intermediate", "Individual", "Accepted for CCA eligibility",
     "cyberab.org"),
    ("8140-612.CISSO", "CISSO", "Mile2", "DoD 8140.3 / 612", "Intermediate",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.Cloud+", "Cloud+", "CompTIA", "DoD 8140.3 / 612",
     "Intermediate", "Individual", "Accepted for CCA eligibility",
     "cyberab.org"),
    ("8140-612.FITSP-A", "FITSP-A", "FITSI", "DoD 8140.3 / 612", "Intermediate",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.GCSA", "GCSA", "GIAC", "DoD 8140.3 / 612", "Intermediate",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.GSEC", "GSEC", "GIAC", "DoD 8140.3 / 612", "Intermediate",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.PenTest+", "PenTest+", "CompTIA", "DoD 8140.3 / 612",
     "Intermediate", "Individual", "Accepted for CCA eligibility",
     "cyberab.org"),
    ("8140-612.Security+", "Security+", "CompTIA", "DoD 8140.3 / 612",
     "Intermediate", "Individual", "Accepted for CCA eligibility",
     "cyberab.org"),
    # 8140.3 / 612 Advanced
    ("8140-612.CCISO", "CCISO", "EC-Council", "DoD 8140.3 / 612", "Advanced",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.CISA", "CISA", "ISACA", "DoD 8140.3 / 612", "Advanced",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.CISM", "CISM", "ISACA", "DoD 8140.3 / 612", "Advanced",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.CISSP", "CISSP", "(ISC)²", "DoD 8140.3 / 612", "Advanced",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.CISSP-ISSEP", "CISSP-ISSEP", "(ISC)²", "DoD 8140.3 / 612",
     "Advanced", "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.CySA+", "CySA+", "CompTIA", "DoD 8140.3 / 612", "Advanced",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.GSLC", "GSLC", "GIAC", "DoD 8140.3 / 612", "Advanced",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    ("8140-612.GSNA", "GSNA", "GIAC", "DoD 8140.3 / 612", "Advanced",
     "Individual", "Accepted for CCA eligibility", "cyberab.org"),
    # SCF Organizational Certifications (issued to Sellers' clients by SCF 3PAOs)
    ("SCF.CORE-Fundamentals", "SCF CORE Fundamentals", "SCF Council",
     "SCF", "—", "Org (issued to OSA)", "Currently available", "cyberab.org"),
    ("SCF.NIST-CSF-2.0", "NIST Cybersecurity Framework 2.0", "SCF Council",
     "SCF", "—", "Org (issued to OSA)", "Currently available", "cyberab.org"),
    ("SCF.NIST-800-161-R1", "NIST SP 800-161 R1 (C-SCRM)", "SCF Council",
     "SCF", "—", "Org (issued to OSA)", "Currently available", "cyberab.org"),
    ("SCF.HIPAA", "HIPAA Security Rule (NIST SP 800-66 R2)", "SCF Council",
     "SCF", "—", "Org (issued to OSA)", "Currently available", "cyberab.org"),
    ("SCF.NYDFS-23NYCRR500", "NY DFS 23 NYCRR 500 (2023 Amendment 2)",
     "SCF Council", "SCF", "—", "Org (issued to OSA)", "Currently available",
     "cyberab.org"),
    ("SCF.NZ-HISF-2022", "New Zealand Health Info Security Framework 2022",
     "SCF Council", "SCF", "—", "Org (issued to OSA)", "Currently available",
     "cyberab.org"),
    ("SCF.CISA-SSDAF", "CISA Secure Software Development Attestation Form",
     "SCF Council", "SCF", "—", "Org (issued to OSA)",
     "Currently available; also CODE 1 for SCA", "cyberab.org"),
    ("SCF.NIST-800-171-R3", "NIST SP 800-171 R3 (non-CMMC)", "SCF Council",
     "SCF", "—", "Org (issued to OSA)", "Currently available", "cyberab.org"),
    ("SCF.FAR-52.204.21", "FAR 52.204.21 (CMMC Level 1)", "SCF Council",
     "SCF", "—", "Org (issued to OSA)", "Currently available", "cyberab.org"),
    ("SCF.CORE-ESP", "SCF CORE External Service Provider (ESP)", "SCF Council",
     "SCF", "—", "Org (issued to OSA)", "Planned 2026", "cyberab.org"),
    ("SCF.AU-Essential8", "Australia Essential Eight", "SCF Council", "SCF",
     "—", "Org (issued to OSA)", "Planned 2026", "cyberab.org"),
    ("SCF.EU-DORA", "EU Digital Operational Resilience Act", "SCF Council",
     "SCF", "—", "Org (issued to OSA)", "Planned 2026", "cyberab.org"),
    ("SCF.EU-NIS2", "ENISA NIS2 (EU 2022/2555)", "SCF Council", "SCF", "—",
     "Org (issued to OSA)", "Planned 2026", "cyberab.org"),
    ("SCF.GLBA", "Gramm-Leach-Bliley Act (16 CFR 314)", "SCF Council", "SCF",
     "—", "Org (issued to OSA)", "Planned 2026", "cyberab.org"),
    # SCA CODE
    ("SCA.CODE1", "CODE Level 1 — CISA SSDAF", "Cyber AB/SCG", "SCA", "—",
     "Org (issued to SDO)", "Maps to EO 14028", "cyberab.org"),
    ("SCA.CODE2", "CODE Level 2 — NIST 800-218", "Cyber AB/SCG", "SCA", "—",
     "Org (issued to SDO)", "SSDF embedded in SDLC", "cyberab.org"),
    ("SCA.CODE3", "CODE Level 3 — Custom", "Cyber AB/SCG", "SCA", "—",
     "Org (issued to SDO)", "Contract/industry-specific", "cyberab.org"),
    # ISO/IEC Accreditation (relevant to Cyber AB chain of trust)
    ("ISO.17011", "ISO/IEC 17011", "ISO/IEC", "Accreditation", "—", "Org",
     "Accreditation bodies (Cyber AB itself)", "cyberab.org + iso.org"),
    ("ISO.17020", "ISO/IEC 17020", "ISO/IEC", "Accreditation", "—", "Org",
     "Inspection bodies (C3PAOs, 3PAOs)", "cyberab.org + iso.org"),
    ("ISO.17024", "ISO/IEC 17024", "ISO/IEC", "Accreditation", "—", "Org",
     "Certification bodies for persons (CAICO, SAICO, IAPP)",
     "cyberab.org + iso.org"),
    # ISO/IEC Management System Lead Auditor tracks
    ("ISO.27001-LA", "ISO/IEC 27001 Lead Auditor", "Multiple (BSI/PECB/etc)",
     "ISO", "Advanced", "Individual", "ISMS", "iso.org + training bodies"),
    ("ISO.27001-LI", "ISO/IEC 27001 Lead Implementer", "Multiple", "ISO",
     "Advanced", "Individual", "ISMS implementation",
     "iso.org + training bodies"),
    ("ISO.27701-LA", "ISO/IEC 27701 Lead Auditor", "Multiple", "ISO",
     "Advanced", "Individual", "Privacy information management",
     "iso.org + training bodies"),
    ("ISO.22301-LA", "ISO 22301 Lead Auditor", "Multiple", "ISO", "Advanced",
     "Individual", "Business continuity", "iso.org + training bodies"),
    ("ISO.9001-LA", "ISO 9001 Lead Auditor", "Multiple", "ISO", "Advanced",
     "Individual", "Quality management", "iso.org + training bodies"),
    ("ISO.42001-LA", "ISO/IEC 42001 Lead Auditor", "Multiple", "ISO",
     "Advanced", "Individual", "AI management",
     "iso.org + training bodies"),
    # (ISC)²
    ("ISC2.CC", "Certified in Cybersecurity", "(ISC)²", "General",
     "Entry", "Individual", "No experience required", "isc2.org"),
    ("ISC2.SSCP", "Systems Security Certified Practitioner", "(ISC)²",
     "General", "Entry", "Individual", "1 yr experience", "isc2.org"),
    ("ISC2.CISSP", "Certified Information Systems Security Professional",
     "(ISC)²", "General", "Advanced", "Individual", "5+ yrs experience",
     "isc2.org"),
    ("ISC2.CCSP", "Certified Cloud Security Professional", "(ISC)²",
     "Cloud", "Advanced", "Individual", "5+ yrs experience", "isc2.org"),
    ("ISC2.CGRC", "Governance, Risk & Compliance (formerly CAP)", "(ISC)²",
     "GRC", "Intermediate", "Individual", "2 yrs experience", "isc2.org"),
    ("ISC2.CSSLP", "Certified Secure Software Lifecycle Professional",
     "(ISC)²", "Secure SDLC", "Advanced", "Individual", "4 yrs experience",
     "isc2.org"),
    ("ISC2.CISSP-ISSAP", "CISSP — Information Security Architecture",
     "(ISC)²", "Architecture", "Senior", "Individual", "CISSP + 2 yrs or 7 yrs",
     "isc2.org"),
    ("ISC2.CISSP-ISSEP", "CISSP — Information Security Engineering",
     "(ISC)²", "Engineering", "Senior", "Individual", "CISSP + 2 yrs or 7 yrs",
     "isc2.org"),
    ("ISC2.CISSP-ISSMP", "CISSP — Information Security Management",
     "(ISC)²", "Management", "Senior", "Individual", "CISSP + 2 yrs or 7 yrs",
     "isc2.org"),
    ("ISC2.HCISPP", "Healthcare Info Security & Privacy Practitioner",
     "(ISC)²", "Healthcare", "Intermediate", "Individual",
     "Legacy — confirm still active", "isc2.org"),
    # ISACA
    ("ISACA.CISA", "Certified Information Systems Auditor", "ISACA", "Audit",
     "Advanced", "Individual", "5 yrs experience; CPE renewal",
     "isaca.org/credentialing"),
    ("ISACA.CISM", "Certified Information Security Manager", "ISACA",
     "Management", "Advanced", "Individual", "5 yrs experience",
     "isaca.org/credentialing"),
    ("ISACA.CRISC", "Certified in Risk and Info Systems Control", "ISACA",
     "Risk", "Advanced", "Individual", "3 yrs experience",
     "isaca.org/credentialing"),
    ("ISACA.CGEIT", "Certified in Governance of Enterprise IT", "ISACA",
     "Governance", "Advanced", "Individual", "5 yrs experience",
     "isaca.org/credentialing"),
    ("ISACA.CDPSE", "Certified Data Privacy Solutions Engineer", "ISACA",
     "Privacy Engineering", "Intermediate", "Individual", "—",
     "isaca.org/credentialing"),
    ("ISACA.CCOA", "Certified Cybersecurity Operations Analyst", "ISACA",
     "SOC Ops", "Intermediate", "Individual", "—",
     "isaca.org/credentialing"),
    ("ISACA.AAIA", "Advanced in AI Audit", "ISACA", "AI Audit", "Advanced",
     "Individual", "—", "isaca.org/credentialing"),
    ("ISACA.AAIR", "Advanced in AI Risk", "ISACA", "AI Risk", "Advanced",
     "Individual", "—", "isaca.org/credentialing"),
    ("ISACA.AAISM", "Advanced in AI Security Management", "ISACA",
     "AI Security", "Advanced", "Individual", "—",
     "isaca.org/credentialing"),
    # CompTIA
    ("CompTIA.A+", "A+", "CompTIA", "IT", "Foundational", "Individual",
     "CEU renewal", "comptia.org"),
    ("CompTIA.Network+", "Network+", "CompTIA", "Networking", "Foundational",
     "Individual", "CEU renewal", "comptia.org"),
    ("CompTIA.Security+", "Security+", "CompTIA", "General", "Entry",
     "Individual", "CEU renewal", "comptia.org"),
    ("CompTIA.CySA+", "Cybersecurity Analyst+", "CompTIA", "Blue team",
     "Intermediate", "Individual", "CEU renewal", "comptia.org"),
    ("CompTIA.PenTest+", "PenTest+", "CompTIA", "Red team", "Intermediate",
     "Individual", "CEU renewal", "comptia.org"),
    ("CompTIA.SecurityX", "SecurityX (formerly CASP+)", "CompTIA", "General",
     "Advanced", "Individual", "CEU renewal", "comptia.org"),
    ("CompTIA.Cloud+", "Cloud+", "CompTIA", "Cloud", "Intermediate",
     "Individual", "CEU renewal", "comptia.org"),
    ("CompTIA.Server+", "Server+", "CompTIA", "Server admin", "Intermediate",
     "Individual", "CEU renewal", "comptia.org"),
    ("CompTIA.Linux+", "Linux+", "CompTIA", "Linux", "Intermediate",
     "Individual", "CEU renewal", "comptia.org"),
    # GIAC (partial — 62 total; picklist may need free-form entry)
    ("GIAC.GSEC", "Security Essentials", "GIAC", "General", "Intermediate",
     "Individual", "4-yr CPE renewal", "giac.org"),
    ("GIAC.GCIH", "Certified Incident Handler", "GIAC", "DFIR",
     "Intermediate", "Individual", "4-yr CPE renewal", "giac.org"),
    ("GIAC.GPEN", "Certified Penetration Tester", "GIAC", "Offensive",
     "Advanced", "Individual", "4-yr CPE renewal", "giac.org"),
    ("GIAC.GWAPT", "Web App Pen Tester", "GIAC", "Offensive", "Advanced",
     "Individual", "4-yr CPE renewal", "giac.org"),
    ("GIAC.GCFA", "Certified Forensic Analyst", "GIAC", "DFIR", "Advanced",
     "Individual", "4-yr CPE renewal", "giac.org"),
    ("GIAC.GSLC", "Security Leadership", "GIAC", "Management", "Advanced",
     "Individual", "4-yr CPE renewal; on DoD 8140", "giac.org"),
    ("GIAC.GSNA", "Systems & Network Auditor", "GIAC", "Audit", "Advanced",
     "Individual", "4-yr CPE renewal; on DoD 8140", "giac.org"),
    ("GIAC.GCSA", "Cloud Security Automation", "GIAC", "Cloud", "Advanced",
     "Individual", "4-yr CPE renewal; on DoD 8140", "giac.org"),
    ("GIAC.GSOA", "Strategic OSINT Analyst", "GIAC", "Intel", "Intermediate",
     "Individual", "—", "giac.org"),
    ("GIAC.GMLE", "Machine Learning Engineer", "GIAC", "AI", "Advanced",
     "Individual", "—", "giac.org"),
    ("GIAC.GRTP", "Red Team Professional", "GIAC", "Offensive", "Advanced",
     "Individual", "—", "giac.org"),
    ("GIAC.GLIR", "Linux Incident Responder", "GIAC", "DFIR", "Intermediate",
     "Individual", "—", "giac.org"),
    # EC-Council
    ("ECC.CEH", "Certified Ethical Hacker", "EC-Council", "Offensive",
     "Intermediate", "Individual", "Widely adopted pentest cert",
     "eccouncil.org"),
    ("ECC.CCISO", "Certified CISO", "EC-Council", "Management", "Senior",
     "Individual", "On DoD 8140", "eccouncil.org"),
    ("ECC.CHFI", "Computer Hacking Forensic Investigator", "EC-Council",
     "DFIR", "Intermediate", "Individual", "—", "eccouncil.org"),
    ("ECC.CND", "Certified Network Defender", "EC-Council", "Defense",
     "Intermediate", "Individual", "—", "eccouncil.org"),
    ("ECC.CASE", "Certified Application Security Engineer", "EC-Council",
     "AppSec", "Intermediate", "Individual", "—", "eccouncil.org"),
    ("ECC.LPT", "Licensed Penetration Tester", "EC-Council", "Offensive",
     "Advanced", "Individual", "—", "eccouncil.org"),
    # Offensive Security (OffSec)
    ("OffSec.OSCP", "Offensive Security Certified Professional", "OffSec",
     "Offensive", "Intermediate", "Individual", "De facto pentester cert",
     "offsec.com"),
    ("OffSec.OSEP", "Experienced Penetration Tester", "OffSec", "Offensive",
     "Advanced", "Individual", "Part of OSCE3", "offsec.com"),
    ("OffSec.OSWE", "Web Expert", "OffSec", "Offensive", "Advanced",
     "Individual", "Part of OSCE3", "offsec.com"),
    ("OffSec.OSED", "Exploit Developer", "OffSec", "Offensive", "Advanced",
     "Individual", "Part of OSCE3", "offsec.com"),
    ("OffSec.OSEE", "Exploitation Expert", "OffSec", "Offensive", "Expert",
     "Individual", "—", "offsec.com"),
    ("OffSec.OSWA", "Web Assessor", "OffSec", "Offensive", "Intermediate",
     "Individual", "—", "offsec.com"),
    ("OffSec.OSMR", "macOS Researcher", "OffSec", "Offensive", "Advanced",
     "Individual", "—", "offsec.com"),
    ("OffSec.OSDA", "Defense Analyst", "OffSec", "Defense", "Advanced",
     "Individual", "—", "offsec.com"),
    ("OffSec.OSCE3", "Certified Expert (OSEP+OSWE+OSED)", "OffSec",
     "Offensive", "Expert", "Individual", "Triple-cert designation",
     "offsec.com"),
    # IAPP Privacy
    ("IAPP.CIPP-E", "CIPP/Europe", "IAPP", "Privacy Law", "Intermediate",
     "Individual", "ANSI ISO 17024 accredited", "iapp.org"),
    ("IAPP.CIPP-US", "CIPP/United States", "IAPP", "Privacy Law",
     "Intermediate", "Individual", "ANSI ISO 17024 accredited", "iapp.org"),
    ("IAPP.CIPP-C", "CIPP/Canada", "IAPP", "Privacy Law", "Intermediate",
     "Individual", "—", "iapp.org"),
    ("IAPP.CIPP-A", "CIPP/Asia", "IAPP", "Privacy Law", "Intermediate",
     "Individual", "—", "iapp.org"),
    ("IAPP.CIPP-China", "CIPP/China", "IAPP", "Privacy Law", "Intermediate",
     "Individual", "PIPL focus", "iapp.org"),
    ("IAPP.CIPM", "Certified Information Privacy Manager", "IAPP",
     "Privacy Mgmt", "Advanced", "Individual", "ANSI ISO 17024 accredited",
     "iapp.org"),
    ("IAPP.CIPT", "Certified Information Privacy Technologist", "IAPP",
     "Privacy Eng", "Intermediate", "Individual",
     "ANSI ISO 17024 accredited", "iapp.org"),
    ("IAPP.AIGP", "AI Governance Professional", "IAPP", "AI Governance",
     "Intermediate", "Individual", "New 2024", "iapp.org"),
    ("IAPP.PLS", "Privacy Law Specialist", "IAPP", "Privacy Law", "Expert",
     "Individual", "Advanced legal credential", "iapp.org"),
    ("IAPP.FIP", "Fellow of Information Privacy", "IAPP", "Privacy",
     "Expert", "Individual", "Requires CIPP + (CIPM or CIPT)", "iapp.org"),
    # CSA Individual
    ("CSA.CCSK", "Cert of Cloud Security Knowledge (v5)", "CSA",
     "Cloud Security", "Intermediate", "Individual", "Foundational cloud",
     "cloudsecurityalliance.org"),
    ("CSA.CCAK", "Cert of Cloud Auditing Knowledge", "CSA + ISACA",
     "Cloud Audit", "Intermediate", "Individual", "Joint with ISACA",
     "cloudsecurityalliance.org"),
    ("CSA.CCZT", "Cert of Competence in Zero Trust", "CSA", "Zero Trust",
     "Intermediate", "Individual", "—", "cloudsecurityalliance.org"),
    # CREST Individual
    ("CREST.CPSA", "CREST Practitioner Security Analyst", "CREST",
     "Offensive", "Intermediate", "Individual", "—", "crest-approved.org"),
    ("CREST.CRT", "CREST Registered Penetration Tester", "CREST",
     "Offensive", "Intermediate", "Individual", "—", "crest-approved.org"),
    ("CREST.CCT-INF", "CREST Certified Tester — Infrastructure", "CREST",
     "Offensive", "Advanced", "Individual", "—", "crest-approved.org"),
    ("CREST.CCT-APP", "CREST Certified Tester — Application", "CREST",
     "Offensive", "Advanced", "Individual", "—", "crest-approved.org"),
    ("CREST.CCT-SAS", "CREST Certified Tester — Simulated Attack", "CREST",
     "Offensive", "Advanced", "Individual", "—", "crest-approved.org"),
    ("CREST.CRIA", "CREST Registered Intrusion Analyst", "CREST", "DFIR",
     "Intermediate", "Individual", "—", "crest-approved.org"),
    ("CREST.CRIS", "CREST Registered Intrusion Specialist", "CREST", "DFIR",
     "Advanced", "Individual", "—", "crest-approved.org"),
    # Mile2
    ("Mile2.CISSO", "Certified Info Systems Security Officer", "Mile2",
     "Management", "Intermediate", "Individual", "On DoD 8140",
     "mile2.com"),
    ("Mile2.CPTE", "Certified Pen Testing Engineer", "Mile2", "Offensive",
     "Intermediate", "Individual", "—", "mile2.com"),
    ("Mile2.CPTC", "Certified Pen Testing Consultant", "Mile2", "Offensive",
     "Advanced", "Individual", "—", "mile2.com"),
    ("Mile2.CDFE", "Certified Digital Forensics Examiner", "Mile2", "DFIR",
     "Intermediate", "Individual", "—", "mile2.com"),
    # FITSI
    ("FITSI.FITSP-A", "FITSP — Auditor", "FITSI", "Federal", "Intermediate",
     "Individual", "On DoD 8140", "fitsi.org"),
    ("FITSI.FITSP-M", "FITSP — Manager", "FITSI", "Federal", "Intermediate",
     "Individual", "—", "fitsi.org"),
    ("FITSI.FITSP-D", "FITSP — Designer", "FITSI", "Federal", "Intermediate",
     "Individual", "—", "fitsi.org"),
    ("FITSI.FITSP-O", "FITSP — Operator", "FITSI", "Federal", "Intermediate",
     "Individual", "—", "fitsi.org"),
]

# ──────────────────────────────────────────────────────────────────────────
# DATA — Role ↔ Cert Requirements (mapping)
# ──────────────────────────────────────────────────────────────────────────

REQUIREMENTS = [
    # (Role ID, Role Name, Requirement Type, Cert ID/Description, Notes)
    ("CMMC.CCP", "CMMC Certified Professional", "Prerequisite",
     "CompTIA.A+ (or equivalent)", "Pre-training recommendation"),
    ("CMMC.CCP", "CMMC Certified Professional", "Prerequisite",
     "DoD CUI Awareness Training", ""),
    ("CMMC.CCP", "CMMC Certified Professional", "Prerequisite",
     "Tier 3 DoD investigation", "Or equivalent clearance"),
    ("CMMC.CCA", "CMMC Certified Assessor", "Prerequisite",
     "Active CMMC.CCP", ""),
    ("CMMC.CCA", "CMMC Certified Assessor", "Prerequisite",
     "3+ yrs cybersecurity experience", ""),
    ("CMMC.CCA", "CMMC Certified Assessor", "Prerequisite",
     "1+ yr assessment/audit experience", ""),
    ("CMMC.CCA", "CMMC Certified Assessor", "Prerequisite",
     "One of 8140-612.* (Intermediate or Advanced)",
     "17 certs accepted — see Certifications sheet"),
    ("CMMC.CCA", "CMMC Certified Assessor", "Prerequisite",
     "US Citizenship + Tier 3/NAC", ""),
    ("CMMC.LCCA", "Lead CMMC Certified Assessor", "Prerequisite",
     "Active CMMC.CCA", ""),
    ("CMMC.LCCA", "Lead CMMC Certified Assessor", "Prerequisite",
     "Enhanced DoD experience", ""),
    ("CMMC.CCI", "CMMC Credentialed Instructor", "Prerequisite",
     "Assessor credentials at level instructed",
     "Launching early 2026"),
    ("CMMC.C3PAO", "CMMC 3PAO (org)", "Requirement",
     "Experian financial review", ""),
    ("CMMC.C3PAO", "CMMC 3PAO (org)", "Requirement",
     "DCSA FOCI review", ""),
    ("CMMC.C3PAO", "CMMC 3PAO (org)", "Requirement",
     "CMMC Level 2 assessment by DCMA DIBCAC", ""),
    ("CMMC.C3PAO", "CMMC 3PAO (org)", "Staffing",
     "Employs CCP/CCA individuals", ""),
    ("SCF.RPO", "SCF RPO (org)", "Staffing",
     "Min 1 SCF.Practitioner OR SCF.Architect", ""),
    ("SCF.APO", "SCF APO (org)", "Staffing",
     "Min 1 SCF.Architect", ""),
    ("SCF.APO", "SCF APO (org)", "Requirement",
     "Completed SCF Platform Assessment", ""),
    ("SCF.APP", "SCF APP (org)", "Staffing",
     "Min 1 SCF.Architect", ""),
    ("SCF.APP", "SCF APP (org)", "Requirement",
     "Integrated SCF into platform", ""),
    ("SCF.3PAO", "SCF 3PAO (org)", "Staffing",
     "Min 2 SCF.Assessor individuals", ""),
    ("SCF.3PAO", "SCF 3PAO (org)", "Requirement",
     "Commercial background + procedural assessment", ""),
    ("SCA.3PAO", "SCA 3PAO (org)", "Requirement",
     "Cyber AB accreditation (typically via SCF 3PAO status)", ""),
    ("FedRAMP.3PAO", "FedRAMP 3PAO (org)", "Requirement",
     "ISO/IEC 17020 accreditation via A2LA", ""),
    ("HITRUST.AEAO", "HITRUST AEAO (org)", "Requirement",
     "Active HITRUST license", ""),
    ("HITRUST.AEAO", "HITRUST AEAO (org)", "Staffing",
     "HITRUST.CCSFP individuals", ""),
    ("PCI.QSA", "PCI QSA (org)", "Staffing",
     "Individually-qualified QSAs via PCI SSC training", ""),
    ("PCI.PFI", "PCI PFI", "Prerequisite",
     "Must work for a PCI.QSA company", ""),
    ("AICPA.CPA-Firm", "Licensed CPA Firm", "Requirement",
     "State CPA license + AICPA peer review", ""),
    ("AICPA.CPA-Firm", "Licensed CPA Firm", "Staffing",
     "AICPA.CPA-Individual + typically ISACA.CISA", "De-facto SOC 2 staffing"),
    ("CSA.STAR-L2-Attestation", "CSA STAR Attestation", "Requirement",
     "Hold SOC 2 report", ""),
    ("CSA.STAR-L2-Certification", "CSA STAR Certification", "Requirement",
     "Hold ISO/IEC 27001 certification", ""),
    ("IAPP.FIP", "IAPP FIP", "Prerequisite",
     "CIPP + (CIPM or CIPT)", ""),
]

# ──────────────────────────────────────────────────────────────────────────
# DATA — Sources (hops done and pending)
# ──────────────────────────────────────────────────────────────────────────

SOURCES = [
    # (URL, Ecosystem(s), Hop Status, Notes)
    ("https://cyberab.org", "CMMC, SCF, SCA", "Done (wget mirror)",
     "195 pages scraped 2026-04-23 via scripts/scrape-cyberab.sh"),
    ("https://isaca.org/cmmc", "CMMC (CAICO admin)", "Done (WebFetch)",
     "Hop 1: confirmed 2025-11-10 launch; 2026-04-01 pricing change; "
     "CCI early 2026"),
    ("https://public.cyber.mil/dcwf-work-role/security-control-assessor/",
     "DoD 8140.3 / 612", "BLOCKED",
     "CAC/SAML-gated via DEAS. Authoritative 8140.3 cert list unreachable "
     "via WebFetch. Fallback: cyberab.org snapshot (Sept 2025)."),
    ("https://fedramp.gov", "FedRAMP", "Partial",
     "Homepage only — full /docs/ pages require deeper crawl"),
    ("https://hitrustalliance.net/assessors/", "HITRUST",
     "Partial (WebFetch)",
     "Confirmed AEAO, CCSFP, CHQP. Full reqs behind login."),
    ("https://pcisecuritystandards.org/assessors_and_solutions/", "PCI",
     "Done (WebFetch)",
     "11 org programs + 3 individual — full list captured"),
    ("https://cloudsecurityalliance.org/star/", "CSA STAR",
     "Done (WebFetch)",
     "STAR L1/L2 + STAR-AI captured. CCSK/CCAK/CCZT added from knowledge."),
    ("https://iapp.org/certify/programs/", "Privacy", "Done (WebFetch)",
     "10 privacy certs incl. new AIGP"),
    ("https://isaca.org/credentialing", "Cross-framework individual",
     "Done (WebFetch)", "9 active + 3 retired credentials"),
    ("https://isc2.org/certifications", "Cross-framework individual",
     "Done (WebFetch)", "9 certs + 3 CISSP concentrations"),
    ("https://comptia.org/certifications", "Cross-framework individual",
     "Partial (WebFetch)",
     "Main cyber stack captured; full catalog pending"),
    ("https://giac.org/certifications", "Cross-framework individual",
     "Partial (WebFetch)",
     "12 of 62 certs visible — full catalog PDF pending"),
    ("https://crest-approved.org/membership/", "CREST",
     "Partial (WebFetch)",
     "Service domains + Pathway tiers; individual certs added from knowledge"),
    ("https://offsec.com/courses", "Cross-framework individual",
     "Not hopped (redirect detected)",
     "Can follow if Clark approves — current list from general knowledge"),
    ("https://eccouncil.org", "Cross-framework individual", "Not hopped",
     "Common EC-Council certs listed from knowledge"),
    ("https://aicpa-cima.com", "SOC 2", "Empty WebFetch",
     "No SOC-specific assessor credential; CPA firm + CISA is practical"),
    ("https://a2la.org", "FedRAMP", "Not hopped",
     "Candidate for next hop — 3PAO accreditation specifics"),
    ("https://nerc.com", "NERC CIP", "Not hopped",
     "Ask before — utility grid sector"),
    ("https://nice.nist.gov", "NICE Framework", "Not hopped",
     "Publicly accessible DoD 8140 superset — candidate substitute "
     "for DCWF hop"),
]

# ──────────────────────────────────────────────────────────────────────────
# Excel writer
# ──────────────────────────────────────────────────────────────────────────

def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E5B88")
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def auto_width(ws, minw=10, maxw=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[letter].width = max(minw, min(maxw, longest + 2))


def wrap_cells(ws, ncols):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook():
    wb = Workbook()

    # Sheet 1 — README
    ws = wb.active
    ws.title = "README"
    readme = [
        ["SME Mart — Compliance Roles & Certifications Catalog"],
        [""],
        ["Built 2026-04-23 from cyberab.org scrape + 1-hop WebFetch across "
         "adjacent compliance ecosystems."],
        [""],
        ["Purpose: Catalog every role and certification that a Buyer in SME "
         "Mart might want a Seller to hold. Seeds a picklist for Seller "
         "profile credentials."],
        [""],
        ["Sheets:"],
        ["  1. Roles — all org and individual ecosystem roles "
         "(CMMC, SCF, SCA, FedRAMP, HITRUST, PCI, SOC 2, CSA, CREST)"],
        ["  2. Certifications — all individual credentials "
         "(8140.3 baselines, ISACA, ISC2, CompTIA, GIAC, OffSec, EC-Council, "
         "IAPP, CSA, CREST, Mile2, FITSI) + SCF org certs + SCA CODE"],
        ["  3. Role-Cert Requirements — mapping (which role requires/confers "
         "which credential)"],
        ["  4. Sources — every URL touched, hop status (Done/Partial/"
         "Blocked/Not hopped), and notes"],
        [""],
        ["Known gaps:"],
        ["  - DoD DCWF 612 authoritative list is CAC-gated; "
         "using cyberab.org Sept-2025 snapshot"],
        ["  - GIAC: only 12 of 62 certs captured via WebFetch; "
         "full catalog PDF pending"],
        ["  - Several adjacent frameworks not hopped (ask before): "
         "NERC CIP, IRAP, ISMAP, C5, Cyber Essentials, StateRAMP, CJIS"],
        [""],
        ["Full narrative + data-model proposals: "
         ".claude/proposals/sme-mart-cyberab-catalog.md (694 lines)"],
        ["Scrape script: scripts/scrape-cyberab.sh"],
        ["Extractor: scripts/extract-cyberab.py"],
        ["This file builder: scripts/build-compliance-catalog-xlsx.py"],
    ]
    for r in readme:
        ws.append(r)
    ws.column_dimensions["A"].width = 110
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws["A1"].font = Font(bold=True, size=14)

    # Sheet 2 — Roles
    ws = wb.create_sheet("Roles")
    ws.append(["ID", "Role", "Ecosystem", "Scope (Org/Individual/Product)",
               "Summary", "Prerequisites", "Fee / Renewal", "Source"])
    for row in ROLES:
        ws.append(list(row))
    style_header(ws, 8)
    wrap_cells(ws, 8)
    auto_width(ws)

    # Sheet 3 — Certifications
    ws = wb.create_sheet("Certifications")
    ws.append(["ID", "Certification", "Issuer", "Ecosystem / Focus",
               "Proficiency / Level", "Scope", "Notes", "Source"])
    for row in CERTS:
        ws.append(list(row))
    style_header(ws, 8)
    wrap_cells(ws, 8)
    auto_width(ws)

    # Sheet 4 — Role-Cert Requirements
    ws = wb.create_sheet("Role-Cert Requirements")
    ws.append(["Role ID", "Role Name", "Requirement Type",
               "Cert / Description", "Notes"])
    for row in REQUIREMENTS:
        ws.append(list(row))
    style_header(ws, 5)
    wrap_cells(ws, 5)
    auto_width(ws)

    # Sheet 5 — Sources
    ws = wb.create_sheet("Sources")
    ws.append(["URL", "Ecosystem(s)", "Hop Status", "Notes"])
    for row in SOURCES:
        ws.append(list(row))
    style_header(ws, 4)
    wrap_cells(ws, 4)
    auto_width(ws, maxw=80)

    return wb


if __name__ == "__main__":
    wb = build_workbook()
    wb.save(OUT)
    # Report
    print(f"Wrote: {OUT}")
    print(f"  Roles: {len(ROLES)}")
    print(f"  Certifications: {len(CERTS)}")
    print(f"  Role-Cert Requirements: {len(REQUIREMENTS)}")
    print(f"  Sources: {len(SOURCES)}")
